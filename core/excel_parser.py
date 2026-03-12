"""
S4 Graficador — Excel Parser
Parsea el Excel de encuestas (3 pestañas) y retorna la estructura de datos.
"""

import openpyxl
from io import BytesIO
from typing import List, Tuple, Dict, Optional
from .models import Answer, Question, SegmentGroup, ParseResult, DEFAULT_PALETTE


def parse_excel(file_bytes: bytes) -> ParseResult:
    """
    Parsea el Excel completo y retorna un ParseResult.

    El Excel tiene 3 pestañas útiles:
    - datos: filas de encuesta con Total + segmentos
    - meta_preguntas: metadata por pregunta (tipo_slide, capítulo, títulos, etc.)
    - meta_respuestas: tabla de colores por label de respuesta
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    warnings = []

    # 1. Leer meta_respuestas → dict {label: color_hex}
    meta_colores = _read_meta_respuestas(wb["meta_respuestas"], warnings)

    # 2. Leer meta_preguntas → dict {id_pregunta: {campos...}}
    meta_preguntas = _read_meta_preguntas(wb["meta_preguntas"], warnings)

    # 3. Parsear pestaña datos
    ws = wb["datos"]

    # 3a. Detectar grupos de segmentos (fila 1) y columnas de segmentos (fila 2)
    segment_cols, segment_groups = _detect_segments(ws)

    # 3b. Detectar bloques de preguntas
    bloques = _detect_blocks(ws, warnings)

    # 3c. Construir Question por bloque
    questions = []
    color_auto_count = 0
    palette_idx = 0

    for start, end, id_preg, preg_texto in bloques:
        if id_preg not in meta_preguntas:
            warnings.append(f"Pregunta '{id_preg}' en datos pero no en meta_preguntas — omitida")
            continue

        meta = meta_preguntas[id_preg]
        respuestas = []

        for row in range(start, end + 1):
            label_raw = ws.cell(row, 3).value  # col C = Respuestas
            if label_raw is None:
                continue
            label = str(label_raw).strip()

            total_raw = ws.cell(row, 4).value  # col D = Total
            total = _safe_float(total_raw)

            seg_vals = [_safe_float(ws.cell(row, c).value) for c, _ in segment_cols]

            # Resolver color
            color = meta_colores.get(label)
            color_pending = color is None
            if color is None:
                color = DEFAULT_PALETTE[palette_idx % len(DEFAULT_PALETTE)]
                palette_idx += 1
                color_auto_count += 1

            respuestas.append(Answer(
                label=label,
                color=color,
                color_pending=color_pending,
                total=total,
                segment_values=seg_vals,
            ))

        if not respuestas:
            warnings.append(f"Pregunta '{id_preg}' sin respuestas válidas — omitida")
            continue

        # Orientación: leer de meta, default "V" (vertical)
        orientacion = str(meta.get("orientacion", "V")).strip().upper()
        if orientacion not in ("H", "V"):
            orientacion = "V"

        questions.append(Question(
            id_pregunta=id_preg,
            tipo_slide=meta["tipo_slide"],
            orientacion=orientacion,
            capitulo=meta["capitulo"],
            cap_app=meta.get("cap_app"),
            titulo_frec=meta["titulo_frec"],
            titulo_app=meta["titulo_app"],
            grupo_frec=int(meta["grupo_frec"]),
            multiple=str(meta.get("multiple", "no")).lower().strip() == "si",
            pregunta_texto=preg_texto or "",
            respuestas=respuestas,
            segment_labels=[label for _, label in segment_cols],
        ))

    return ParseResult(
        questions=questions,
        segment_labels=[label for _, label in segment_cols],
        segment_groups=segment_groups,
        warnings=warnings,
        color_auto_assigned=color_auto_count,
    )


def _read_meta_respuestas(ws, warnings: list) -> Dict[str, str]:
    """Lee la pestaña meta_respuestas y retorna {label: color_hex}."""
    colores = {}
    for row in range(2, ws.max_row + 1):
        label_raw = ws.cell(row, 1).value
        if label_raw is None:
            continue
        label = str(label_raw).strip()
        color_raw = ws.cell(row, 2).value
        if color_raw is not None:
            color = str(color_raw).strip()
            if not color.startswith("#"):
                color = f"#{color}"
            colores[label] = color
    return colores


def _read_meta_preguntas(ws, warnings: list) -> Dict[str, dict]:
    """Lee la pestaña meta_preguntas y retorna {id_pregunta: {campos}}."""
    # Leer headers de fila 1
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val is not None:
            headers.append((col, str(val).strip()))
        else:
            break

    if not headers:
        warnings.append("meta_preguntas: no se encontraron headers")
        return {}

    # Mapear nombre de columna a índice
    col_map = {}
    for col_idx, name in headers:
        # Normalizar nombres
        name_lower = name.lower().strip()
        col_map[name_lower] = col_idx
        col_map[name] = col_idx  # También guardar original

    meta = {}
    for row in range(2, ws.max_row + 1):
        id_preg_raw = ws.cell(row, 1).value
        if id_preg_raw is None:
            continue
        id_preg = str(id_preg_raw).strip()

        entry = {
            "id_pregunta": id_preg,
            "tipo_slide": _get_cell_str(ws, row, col_map, "tipo_slide", "FREC_SIMPLE"),
            "capitulo": _get_cell_str(ws, row, col_map, "capitulo", "Sin capítulo"),
            "titulo_frec": _get_cell_str(ws, row, col_map, "titulo_frec", id_preg),
            "titulo_app": _get_cell_str(ws, row, col_map, "titulo_app", id_preg),
            "grupo_frec": _get_cell_val(ws, row, col_map, "grupo_frec", 0),
            "multiple": _get_cell_str(ws, row, col_map, "multiple", "no"),
            "orientacion": _get_cell_str(ws, row, col_map, "orientacion", "V"),
        }

        # CAP_APP puede ser None (la apertura va después de la frecuencia)
        cap_app_col = col_map.get("cap_app") or col_map.get("CAP_APP")
        if cap_app_col:
            cap_app_val = ws.cell(row, cap_app_col).value
            if cap_app_val is not None and str(cap_app_val).strip():
                entry["cap_app"] = str(cap_app_val).strip()
            else:
                entry["cap_app"] = None
        else:
            entry["cap_app"] = None

        meta[id_preg] = entry

    return meta


def _detect_segments(ws) -> Tuple[List[Tuple[int, str]], List[SegmentGroup]]:
    """
    Detecta columnas de segmentos válidas (fila 2, cols F+, solo las no-vacías)
    y agrupa por agrupadores de fila 1.
    """
    segment_cols = []  # [(col_index, label)]
    segment_groups = []  # [SegmentGroup]

    current_group_name = None
    current_group_labels = []

    for col in range(6, ws.max_column + 1):  # Desde col F (index 6)
        header_val = ws.cell(2, col).value
        group_val = ws.cell(1, col).value

        # Si fila 1 tiene valor, es un nuevo grupo
        if group_val is not None and str(group_val).strip():
            # Guardar grupo anterior
            if current_group_name and current_group_labels:
                segment_groups.append(SegmentGroup(
                    name=current_group_name,
                    labels=current_group_labels,
                ))
            current_group_name = str(group_val).strip()
            current_group_labels = []

        # Solo tomar columnas con header no vacío
        if header_val is not None and str(header_val).strip():
            label = str(header_val).strip()
            segment_cols.append((col, label))
            current_group_labels.append(label)

    # Último grupo
    if current_group_name and current_group_labels:
        segment_groups.append(SegmentGroup(
            name=current_group_name,
            labels=current_group_labels,
        ))

    return segment_cols, segment_groups


def _detect_blocks(ws, warnings: list) -> List[Tuple[int, int, str, str]]:
    """
    Detecta bloques de preguntas en la pestaña datos.
    Un bloque empieza cuando col A tiene un valor distinto al anterior.
    Con data_only=True, las fórmulas =+A3 se resuelven al valor cacheado.

    Retorna: [(start_row, end_row, id_pregunta, pregunta_texto)]
    """
    blocks = []
    current_id = None
    current_start = None
    current_preg = None

    for row in range(3, ws.max_row + 1):  # Datos desde fila 3
        id_val = ws.cell(row, 1).value  # col A
        preg_val = ws.cell(row, 2).value  # col B

        # Fila completamente vacía (sin id ni respuesta) → fin de datos
        resp_val = ws.cell(row, 3).value
        if id_val is None and resp_val is None:
            # Podría ser una fila vacía intermedia, o fin real
            # Verificar si hay más datos adelante (check next 3 rows)
            has_more = False
            for ahead in range(1, 4):
                if row + ahead <= ws.max_row:
                    if ws.cell(row + ahead, 1).value is not None or ws.cell(row + ahead, 3).value is not None:
                        has_more = True
                        break
            if not has_more:
                break
            continue

        if id_val is None:
            continue

        id_str = str(id_val).strip()
        preg_str = str(preg_val).strip() if preg_val else ""

        if id_str != current_id:
            # Nuevo bloque
            if current_id is not None:
                blocks.append((current_start, row - 1, current_id, current_preg))
            current_id = id_str
            current_start = row
            current_preg = preg_str

    # Último bloque
    if current_id is not None:
        # Encontrar la última fila real del bloque
        last_row = current_start
        for r in range(current_start, ws.max_row + 1):
            if ws.cell(r, 3).value is not None:
                last_row = r
            elif ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip() == current_id:
                if ws.cell(r, 3).value is not None:
                    last_row = r
            else:
                break
        blocks.append((current_start, last_row, current_id, current_preg))

    return blocks


def _safe_float(val) -> float:
    """Convierte un valor a float de forma segura."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _get_cell_str(ws, row: int, col_map: dict, name: str, default: str) -> str:
    """Lee una celda por nombre de columna y retorna string."""
    col = col_map.get(name) or col_map.get(name.lower())
    if col is None:
        return default
    val = ws.cell(row, col).value
    if val is None:
        return default
    return str(val).strip()


def _get_cell_val(ws, row: int, col_map: dict, name: str, default):
    """Lee una celda por nombre de columna y retorna el valor raw."""
    col = col_map.get(name) or col_map.get(name.lower())
    if col is None:
        return default
    val = ws.cell(row, col).value
    if val is None:
        return default
    return val
